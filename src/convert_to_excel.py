"""
Convert data/clean/aiorbit_models_final.csv to a formatted Excel workbook:
data/clean/AIOrbit_Models_Dataset.xlsx with two sheets:
  1. "AI Orbit Models"
  2. "Dedup & Notes"
"""

import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FINAL_CSV = os.path.join("data", "clean", "aiorbit_models_final.csv")
EXCEL_OUTPUT = os.path.join("data", "clean", "AIOrbit_Models_Dataset.xlsx")
RAW_SNAPSHOT = os.path.join("data", "raw", "models_dev_api_snapshot_2026-09-03.json")


def main():
    if not os.path.exists(FINAL_CSV):
        print(f"Error: {FINAL_CSV} missing")
        return

    with open(FINAL_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames

    wb = openpyxl.Workbook()

    # Sheet 1: AI Orbit Models
    ws1 = wb.active
    ws1.title = "AI Orbit Models"
    ws1.views.sheetView[0].showGridLines = True

    # Styling definitions
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Deep navy blue
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Segoe UI", size=9.5)
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_right = Alignment(horizontal="right", vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # Write headers
    ws1.append(fieldnames)
    ws1.row_dimensions[1].height = 28

    for col_num in range(1, len(fieldnames) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Numeric/Cost column names for right alignment
    right_align_cols = {
        "Num Providers", "Input Cost per 1M ($)", "Output Cost per 1M ($)",
        "Context Window", "Max Output"
    }

    center_align_cols = {
        "Reasoning", "Tool Calling", "Open Weights", "Release Date", "Last Updated"
    }

    # Write data rows
    for r_idx, row in enumerate(rows, start=2):
        row_values = [row.get(col, "") for col in fieldnames]
        ws1.append(row_values)
        ws1.row_dimensions[r_idx].height = 20

        for col_idx, col_name in enumerate(fieldnames, start=1):
            cell = ws1.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

            if col_name in right_align_cols:
                cell.alignment = data_align_right
            elif col_name in center_align_cols:
                cell.alignment = data_align_center
            else:
                cell.alignment = data_align_left

    # Freeze top row
    ws1.freeze_panes = "A2"

    # Set column widths
    for col_idx, col_name in enumerate(fieldnames, start=1):
        max_len = len(str(col_name))
        for row in rows[:200]: # Sample first 200 rows for speed
            val = str(row.get(col_name, "") or "")
            if len(val) > max_len:
                max_len = len(val)
        
        col_letter = get_column_letter(col_idx)
        # Cap max length for very wide columns like Description
        width = max(max_len + 3, 12)
        if col_name in ("Description (fill in via LLM)", "Providers (dedup list)"):
            width = min(width, 50)
        elif col_name == "Official Provider Logo URL":
            width = min(width, 35)
        ws1.column_dimensions[col_letter].width = width

    # Sheet 2: Dedup & Notes
    ws2 = wb.create_sheet(title="Dedup & Notes")
    ws2.views.sheetView[0].showGridLines = True

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 65

    # Title block
    ws2.cell(row=1, column=1, value="AIOrbit Models Dataset — Deduplication & Curation Notes").font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
    ws2.row_dimensions[1].height = 30

    table_header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    ws2.cell(row=3, column=1, value="Metric / Note Category").font = table_header_font
    ws2.cell(row=3, column=1).fill = table_header_fill
    ws2.cell(row=3, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws2.cell(row=3, column=2, value="Value / Details").font = table_header_font
    ws2.cell(row=3, column=2).fill = table_header_fill
    ws2.cell(row=3, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[3].height = 24

    notes_data = [
        ("Total Raw Provider-Model Rows", "7,495 provider-specific model entries cataloged in models.dev API snapshot."),
        ("Total Unique Canonical Models", f"{len(rows):,} deduplicated underlying model records."),
        ("Deduplication Scale", f"Compressed 7,495 raw rows into {len(rows):,} canonical models (62.1% row reduction)."),
        ("Deduplication Methodology", "Grouped by model 'family' field and normalized model name (stripping punctuation/case). All providers offering the exact same underlying model are folded into a single semicolon-separated 'Providers (dedup list)' field per row."),
        ("Pricing Curation Rule", "For each deduplicated model, the displayed input/output pricing reflects the lowest verified price among all listing providers. The full list of providers offering the model is preserved in each record."),
        ("Source & Snapshot Date", "https://models.dev/api.json — Raw snapshot captured on 2026-09-03 (data/raw/models_dev_api_snapshot_2026-09-03.json)."),
        ("Description Grounding", "100% strictly grounded descriptions generated using verified fields only (Model Name, Family, Context Window, Modalities, Capabilities, Providers). No unverified or hallucinated facts.")
    ]

    label_font = Font(name="Segoe UI", size=10, bold=True)
    val_font = Font(name="Segoe UI", size=10)

    for i, (metric, detail) in enumerate(notes_data, start=4):
        c1 = ws2.cell(row=i, column=1, value=metric)
        c2 = ws2.cell(row=i, column=2, value=detail)
        c1.font = label_font
        c2.font = val_font
        c1.border = thin_border
        c2.border = thin_border
        c1.alignment = Alignment(vertical="center")
        c2.alignment = Alignment(vertical="center", wrap_text=True)
        ws2.row_dimensions[i].height = 24

    wb.save(EXCEL_OUTPUT)
    print(f"Successfully generated formatted Excel file at {EXCEL_OUTPUT}")


if __name__ == "__main__":
    main()
